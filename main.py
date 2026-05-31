import asyncio
import re
import smtplib
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from enum import IntEnum, auto
from io import BytesIO
from urllib.parse import urlparse

import openpyxl
from environs import Env
from playwright.async_api import Page, async_playwright

env = Env()
env.read_env()

RE_404 = re.compile(r'\b404\b')
RE_SPACES = re.compile(r'\s+')
RE_ERRO = re.compile(r'not found|nao encontrada|nao encontrado|erro|manutencao|offline')
RE_BOT = re.compile(r'cloudflare|captcha|attention required|checking your browser|robot|sou humano|permission')
RE_ESGOTADO = re.compile(r'esgotado|indisponivel|nao ha ingressos|encerrado')
RE_ATIVO = re.compile(r'comprar|ingresso|selecionar assento|checkout|entrada')


@dataclass(frozen=True)
class Config:
    SPREADSHEET_URL = env.str('SPREADSHEET_URL')
    SHEET_NAME = env.str('SHEET_NAME', default='')
    LINK_COLUMN_NAME = env.str('LINK_COLUMN_NAME')
    DATE_COLUMN_NAME = env.str('DATE_COLUMN_NAME', default='')
    PAGE_LOAD_TIMEOUT = env.int('PAGE_LOAD_TIMEOUT', default=30000)
    MAX_CONCURRENT_PAGES = env.int('MAX_CONCURRENT_PAGES', default=2)

    SEND_EMAIL = env.bool('SEND_EMAIL', default=True)
    SMTP_HOST = env.str('SMTP_HOST')
    SMTP_PORT = env.int('SMTP_PORT', default=587)
    SMTP_USER = env.str('SMTP_USER', default='')
    SMTP_PASSWORD = env.str('SMTP_PASSWORD', default='')
    EMAIL_FROM = env.str('EMAIL_FROM', default='')
    EMAIL_TO = env.str('EMAIL_TO')


class Status(IntEnum):
    OK = auto()
    ERRO_REDE = auto()
    ERRO_SERVIDOR = auto()
    ERRO_CLIENTE = auto()
    ERRO_EXECUCAO = auto()
    BLOQUEADO_BOT = auto()
    PAGINA_EM_BRANCO = auto()
    PAGINA_MENSAGEM_ERRO = auto()
    PAGINA_MENSAGEM_ESGOTADO = auto()


@dataclass(frozen=True)
class LinkResult:
    url: str
    status: Status
    context: str = ''

    def __str__(self) -> str:
        return f'{self.url} - {self.status.name}: {self.context}'


async def get_sheet_links() -> list[tuple[int, str]]:
    url = Config.SPREADSHEET_URL.split('/edit')[0] + '/export?format=xlsx'

    async with async_playwright() as p:
        request_context = await p.request.new_context()
        response = await request_context.get(url)
        if not response.ok:
            raise RuntimeError(f'Fetch failed: {response.status}')
        body = await response.body()
        wb = openpyxl.load_workbook(BytesIO(body), data_only=False)

    sheet = None
    if Config.SHEET_NAME:
        try:
            sheet_idx = wb.sheetnames.index(Config.SHEET_NAME)
        except ValueError:
            print(f'Sheet named {Config.SHEET_NAME} does not exst in spreasheet. Defaulting to first active one')
        else:
            sheet = wb.worksheets[sheet_idx]

    if sheet is None:
        sheet = wb.active
    assert sheet is not None

    link_col_idx = None
    date_col_idx = None
    for link_cell in sheet[1]:
        match link_cell.value:
            case Config.LINK_COLUMN_NAME:
                link_col_idx = link_cell.column
            case Config.DATE_COLUMN_NAME:
                date_col_idx = link_cell.column

    if not link_col_idx:
        raise ValueError(f'Link column {Config.LINK_COLUMN_NAME} not found')
    if Config.DATE_COLUMN_NAME and not date_col_idx:
        raise ValueError(f'Date column {Config.DATE_COLUMN_NAME} not found')

    links = []
    for row in range(2, sheet.max_row + 1):
        if date_col_idx:
            date_cell = sheet.cell(row=row, column=date_col_idx)
            date = date_cell.value

            if not (isinstance(date, datetime) and date > datetime.now()):
                continue

        link_cell = sheet.cell(row=row, column=link_col_idx)
        if link_cell.hyperlink and link_cell.hyperlink.target:
            links.append((row, link_cell.hyperlink.target))

    return links


def clean_text(text: str) -> str:
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
    text = RE_SPACES.sub(' ', text)
    return text.lower().strip()


async def get_text(page: Page) -> str:
    script = """
() => {
    let text = [];
    function walk(node) {
        if (node.nodeType === Node.TEXT_NODE) {
            const content = node.textContent.trim();
            if (content) text.push(content);
            return;
        }

        if (node.nodeType === Node.ELEMENT_NODE) {
            const tag = node.tagName;
            if (tag === 'SCRIPT' || tag === 'STYLE' || tag === 'NOSCRIPT') return;
            if (node.shadowRoot) walk(node.shadowRoot);
        }

        for (const child of node.childNodes) walk(child);
    }

    walk(document.body);
    return text.join(' ');
}
"""

    try:
        return await page.evaluate(script)
    except Exception:
        return ''


async def check_link(page: Page, url: str) -> LinkResult:
    try:
        response = await page.goto(url, wait_until='networkidle', timeout=Config.PAGE_LOAD_TIMEOUT)

        if not response:
            return LinkResult(url, Status.ERRO_REDE)

        if response.status >= 500:
            return LinkResult(url, Status.ERRO_SERVIDOR)
        if response.status >= 400:
            return LinkResult(url, Status.ERRO_CLIENTE)

        titulo = clean_text(await page.title())
        if RE_ERRO.search(titulo) or RE_404.search(titulo):
            return LinkResult(url, Status.PAGINA_MENSAGEM_ERRO)
        if RE_BOT.search(titulo):
            return LinkResult(url, Status.BLOQUEADO_BOT)

        raw_text = await get_text(page)
        texto_corpo = clean_text(raw_text)

        if not texto_corpo:
            print(raw_text)
            return LinkResult(url, Status.PAGINA_EM_BRANCO)

        if RE_ERRO.search(texto_corpo) or RE_404.search(texto_corpo):
            return LinkResult(url, Status.PAGINA_MENSAGEM_ERRO)

        if RE_BOT.search(texto_corpo):
            return LinkResult(url, Status.BLOQUEADO_BOT)

        if RE_ESGOTADO.search(texto_corpo) and not RE_ATIVO.search(texto_corpo):
            return LinkResult(url, Status.PAGINA_MENSAGEM_ESGOTADO)

        return LinkResult(url, Status.OK)
    except Exception as e:
        return LinkResult(url, Status.ERRO_EXECUCAO, str(e))


def get_root_domain(url):
    netloc = urlparse(url).netloc.split(':')[0]
    return next((p for p in reversed(netloc.split('.')) if len(p) > 3), netloc)


async def process_row(sem: asyncio.Semaphore, context, row, link):
    async with sem:
        page = await context.new_page()
        await page.route(
            '**/*',
            lambda route: (
                route.abort()
                if route.request.resource_type in ['image', 'stylesheet', 'media', 'font', 'other', 'manifest']
                or (
                    route.request.resource_type in ['script', 'fetch', 'xhr']
                    and get_root_domain(route.request.url) != get_root_domain(link)
                )
                else route.continue_()
            ),
        )
        res = await check_link(page, link)
        await page.close()

        if res.status != Status.OK:
            print(row, res)
            return (row, res)
        return None


def send_error_report(failures: list[tuple[int, LinkResult]], total_links: int):
    if not Config.SMTP_HOST or not Config.EMAIL_TO:
        print('SMTP config missing')
        return

    fail_count = len(failures)

    msg = EmailMessage()
    msg['Subject'] = f'[{fail_count}/{total_links} Falhas] Monitoramento de Links'
    email_from = Config.EMAIL_FROM if Config.EMAIL_FROM else Config.SMTP_USER
    msg['From'] = email_from
    msg['To'] = Config.EMAIL_TO
    body = ''
    for row, res in failures:
        body += f'\n[Linha {row:03d}] {res.status.name}'
        body += f'\nURL:    {res.url}'
        if res.context:
            body += f'\nMotivo: {res.context}\n'

    msg.set_content(body)

    try:
        with smtplib.SMTP(Config.SMTP_HOST, Config.SMTP_PORT) as server:
            server.starttls()
            server.login(Config.SMTP_USER, Config.SMTP_PASSWORD)
            server.send_message(msg)
        print('Email enviado.')
    except Exception as e:
        print(f'Erro SMTP: {e!s}')


async def main():
    links = await get_sheet_links()
    total_links = len(links)
    if total_links == 0:
        return

    sem = asyncio.Semaphore(Config.MAX_CONCURRENT_PAGES)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                '--disable-gpu',
                '--disable-dev-shm-usage',
                '--disable-extensions',
                '--no-sandbox',
                '--disable-setuid-sandbox',
                '--disable-gl-drawing-for-tests',
                '--disable-animations',
                '--blink-settings=imagesEnabled=false',
                '--mute-audio',
                '--hide-scrollbars',
                '--disable-site-isolation-trials',
                '--disable-webgl',
                '--disable-background-networking',
                '--single-process',
            ],
        )
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
            extra_http_headers={'sec-ch-ua': '"Chromium";v="125", "Not.A/Brand";v="24"'},
        )

        tasks = [process_row(sem, context, row, url) for row, url in links]
        results = await asyncio.gather(*tasks)

        await browser.close()
    failures = [r for r in results if r is not None]

    if failures:
        print(f'Foram encontrados {len(failures)}/{total_links} links com erro')
        if Config.SEND_EMAIL:
            send_error_report(failures, total_links)
    else:
        print(f'Sucesso. {total_links} links checados. Zero falhas.')


if __name__ == '__main__':
    asyncio.run(main())
