from datetime import date, datetime, time
from io import BytesIO
from typing import Any

import openpyxl

from config import (
    CheckConfig,
    DateFilter,
    FilterConfig,
    MatchColumnCheck,
    MatchTextCheck,
    NumberFilter,
    TextFilter,
)
from shared import OPERATOR_MAP, Assertion, CompiledCheck, ScrapeTask
from utils import clean_text


def _evaluate_filter(raw_cell_value: Any, filter: FilterConfig) -> bool:
    if raw_cell_value is None:
        return False

    try:
        match filter:
            case NumberFilter():
                cell_val = float(raw_cell_value)
                target_val = filter.value

            case TextFilter():
                cell_val = clean_text(str(raw_cell_value))
                target_val = filter.value

            case DateFilter():
                val_str = filter.value

                if val_str == 'now':
                    target_val = datetime.now()
                elif val_str == 'today':
                    target_val = datetime.combine(date.today(), time.min)
                elif 't' in val_str.lower() or ':' in val_str:
                    target_val = datetime.fromisoformat(val_str)
                else:
                    target_val = datetime.combine(date.fromisoformat(val_str), time.min)

                if isinstance(raw_cell_value, datetime):
                    cell_val = raw_cell_value
                else:
                    if 't' in str(raw_cell_value).lower() or ':' in str(raw_cell_value):
                        cell_val = datetime.fromisoformat(raw_cell_value)
                    else:
                        cell_val = datetime.combine(date.fromisoformat(raw_cell_value), time.min)

        op_fn = OPERATOR_MAP[filter.operator]
        return op_fn(cell_val, target_val)

    except (ValueError, TypeError) as e:
        print(e)
        return False


def get_scrapper_tasks(
    bytes: BytesIO,
    sheetname: str | None,
    link_column: str,
    filters: list[FilterConfig],
    checks_to_compile: list[CheckConfig],
) -> list[ScrapeTask]:
    wb = openpyxl.load_workbook(bytes, data_only=True)

    sheet = None
    if (target_name := sheetname) is not None:
        for idx, name in enumerate(wb.sheetnames):
            if clean_text(name) == target_name:
                sheet = wb.worksheets[idx]
                break

    if sheet is None:
        sheet = wb.active
    assert sheet is not None

    needed_cols = {link_column}
    for filter in filters:
        needed_cols.add(filter.column)
    for check in checks_to_compile:
        match check:
            case MatchTextCheck():
                assert check.is_conditional

                needed_cols.add(check.if_column)  # pyright: ignore[reportArgumentType]
            case MatchColumnCheck():
                needed_cols.add(check.column)
    col_map = {clean_text(str(cell.value)): idx + 1 for idx, cell in enumerate(sheet[1])}

    tasks = []
    for row in range(2, sheet.max_row + 1):
        link_cell = sheet.cell(row=row, column=col_map[link_column])
        hyperlink = link_cell.hyperlink
        if not hyperlink or not hyperlink.target:
            continue

        row_data = {col: sheet.cell(row=row, column=col_map[col]).value for col in needed_cols}
        keep_row = True
        for filter in filters:
            keep_row = _evaluate_filter(row_data[filter.column], filter)
            if not keep_row:
                break

        if not keep_row:
            continue

        compiled = []
        for check in checks_to_compile:
            match check:
                case MatchTextCheck():
                    meet_condition = OPERATOR_MAP[check.if_operator](check.if_value, row_data[check.if_column])  # pyright: ignore[reportArgumentType]
                    if not meet_condition:
                        continue

                    compiled.append(
                        CompiledCheck(
                            check.error,
                            targets=check.text,  # pyright: ignore[reportArgumentType]
                            assertion=check.assertion,  # pyright: ignore[reportArgumentType]
                            screenshot=check.screenshot,
                        )
                    )

                case MatchColumnCheck():
                    target = row_data[check.column]
                    if not isinstance(target, datetime):
                        target = clean_text(str(target))
                    compiled.append(
                        CompiledCheck(
                            error=check.error,
                            targets=[target],  # pyright: ignore[reportArgumentType]
                            assertion=Assertion.REQUIRED,
                            screenshot=check.screenshot,
                        )
                    )
        tasks.append(ScrapeTask(row=row, url=hyperlink.target, checks=compiled))

    return tasks
